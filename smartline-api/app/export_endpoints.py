"""
Export & Reporting Endpoints - Phase 4B
========================================
"""

from fastapi import APIRouter, Query, Depends, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional, Dict, Any
from datetime import datetime
from decimal import Decimal
import csv
import io
import os
import psycopg2
from psycopg2.extras import RealDictCursor
import traceback

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from io import BytesIO

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.pdfgen import canvas

def get_db():
    """Database connection dependency."""
    conn = psycopg2.connect(
        host=os.environ["PGHOST"],
        dbname=os.environ["PGDATABASE"],
        user=os.environ["PGUSER"],
        password=os.environ["PGPASSWORD"],
        port=os.environ.get("PGPORT", 5432),
        cursor_factory=RealDictCursor
    )
    try:
        yield conn
    finally:
        conn.close()


router = APIRouter(prefix="/bankroll/export", tags=["Export & Reports"])


@router.get("/csv")
async def export_csv(
    user_id: int = Query(default=1),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    bookmaker: Optional[str] = Query(default=None),
    market: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    search: Optional[str] = Query(default=None),
    conn = Depends(get_db)
):
    """Export bets to CSV format with optional filtering."""
    cursor = None
    try:
        cursor = conn.cursor()
        
        # Build dynamic WHERE clause
        where_clauses = ["b.user_id = %s"]
        params = [user_id]
        
        if start_date:
            where_clauses.append("DATE(b.placed_at) >= %s")
            params.append(start_date)
        
        if end_date:
            where_clauses.append("DATE(b.placed_at) <= %s")
            params.append(end_date)
        
        if bookmaker:
            where_clauses.append("ba.bookmaker_name = %s")
            params.append(bookmaker)
        
        if market:
            where_clauses.append("b.market_key = %s")
            params.append(market)
        
        if status:
            where_clauses.append("b.status = %s")
            params.append(status)
        
        if search:
            where_clauses.append("(b.notes ILIKE %s OR b.bet_side ILIKE %s)")
            search_pattern = f"%{search}%"
            params.extend([search_pattern, search_pattern])
        
        where_clause = " AND ".join(where_clauses)
        
        # Fetch bets with all details
        query = f"""
            SELECT 
                b.bet_id,
                TO_CHAR(b.placed_at, 'YYYY-MM-DD HH24:MI:SS') as placed_at,
                ba.bookmaker_name,
                b.bet_type,
                b.sport,
                b.market_key,
                b.bet_side,
                b.line_value,
                b.odds_american,
                b.stake_amount,
                b.potential_payout,
                b.status,
                b.actual_payout,
                b.profit_loss,
                TO_CHAR(b.settled_at, 'YYYY-MM-DD HH24:MI:SS') as settled_at,
                b.notes
            FROM bets b
            LEFT JOIN bankroll_accounts ba ON b.account_id = ba.account_id
            WHERE {where_clause}
            ORDER BY b.placed_at DESC
        """
        
        cursor.execute(query, params)
        bets = cursor.fetchall()
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header (removed Odds Decimal)
        writer.writerow([
            'Bet ID', 'Date Placed', 'Bookmaker', 'Bet Type', 'Sport', 
            'Market', 'Side', 'Line', 'Odds (American)',
            'Stake', 'Potential Payout', 'Status', 'Actual Payout', 
            'Profit/Loss', 'Settled Date', 'Notes'
        ])
        
        # Write data
        for bet in bets:
            writer.writerow([
                bet['bet_id'],
                bet['placed_at'],
                bet['bookmaker_name'] or 'N/A',
                bet['bet_type'],
                bet['sport'],
                bet['market_key'],
                bet['bet_side'],
                bet['line_value'] or '',
                bet['odds_american'] or '',
                float(bet['stake_amount']) if bet['stake_amount'] else '',
                float(bet['potential_payout']) if bet['potential_payout'] else '',
                bet['status'],
                float(bet['actual_payout']) if bet['actual_payout'] else '',
                float(bet['profit_loss']) if bet['profit_loss'] else '',
                bet['settled_at'] or '',
                bet['notes'] or ''
            ])
        
        output.seek(0)
        filename = f"bets_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        if cursor:
            cursor.close()
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        print(f"❌ Export CSV Error: {str(e)}")
        traceback.print_exc()
        if cursor:
            try:
                cursor.close()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/summary")
async def get_export_summary(
    user_id: int = Query(default=1),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    bookmaker: Optional[str] = Query(default=None),
    market: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    conn = Depends(get_db)
):
    """Get summary statistics for export preview."""
    cursor = None
    try:
        cursor = conn.cursor()
        
        where_clauses = ["b.user_id = %s"]
        params = [user_id]
        
        if start_date:
            where_clauses.append("DATE(b.placed_at) >= %s")
            params.append(start_date)
        
        if end_date:
            where_clauses.append("DATE(b.placed_at) <= %s")
            params.append(end_date)
        
        if bookmaker:
            where_clauses.append("ba.bookmaker_name = %s")
            params.append(bookmaker)
        
        if market:
            where_clauses.append("b.market_key = %s")
            params.append(market)
        
        if status:
            where_clauses.append("b.status = %s")
            params.append(status)
        
        where_clause = " AND ".join(where_clauses)
        
        query = f"""
            SELECT 
                COUNT(*) as total_bets,
                COUNT(*) FILTER (WHERE b.status = 'won') as won_bets,
                COUNT(*) FILTER (WHERE b.status = 'lost') as lost_bets,
                COUNT(*) FILTER (WHERE b.status = 'push') as push_bets,
                COUNT(*) FILTER (WHERE b.status = 'pending') as pending_bets,
                COALESCE(SUM(b.stake_amount), 0) as total_staked,
                COALESCE(SUM(b.profit_loss) FILTER (WHERE b.status IN ('won', 'lost', 'push')), 0) as total_profit_loss,
                ROUND(
                    CASE 
                        WHEN COUNT(*) FILTER (WHERE b.status IN ('won', 'lost')) > 0 THEN
                            (COUNT(*) FILTER (WHERE b.status = 'won')::numeric / 
                             COUNT(*) FILTER (WHERE b.status IN ('won', 'lost')) * 100)
                        ELSE 0
                    END,
                1) as win_rate,
                MIN(DATE(b.placed_at)) as earliest_bet,
                MAX(DATE(b.placed_at)) as latest_bet
            FROM bets b
            LEFT JOIN bankroll_accounts ba ON b.account_id = ba.account_id
            WHERE {where_clause}
        """
        
        cursor.execute(query, params)
        summary = cursor.fetchone()
        
        if cursor:
            cursor.close()
        
        return {
            "total_bets": summary['total_bets'],
            "won_bets": summary['won_bets'],
            "lost_bets": summary['lost_bets'],
            "push_bets": summary['push_bets'],
            "pending_bets": summary['pending_bets'],
            "total_staked": float(summary['total_staked']),
            "total_profit_loss": float(summary['total_profit_loss']),
            "win_rate": float(summary['win_rate']),
            "earliest_bet": str(summary['earliest_bet']) if summary['earliest_bet'] else None,
            "latest_bet": str(summary['latest_bet']) if summary['latest_bet'] else None,
            "date_range_days": (summary['latest_bet'] - summary['earliest_bet']).days if summary['earliest_bet'] and summary['latest_bet'] else 0
        }
        
    except Exception as e:
        print(f"❌ Summary Error: {str(e)}")
        traceback.print_exc()
        if cursor:
            try:
                cursor.close()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"Summary failed: {str(e)}")


@router.get("/filter-options")
async def get_filter_options(
    user_id: int = Query(default=1),
    conn = Depends(get_db)
):
    """Get available filter options."""
    cursor = None
    try:
        cursor = conn.cursor()
        
        # Get unique bookmakers
        cursor.execute("""
            SELECT DISTINCT ba.bookmaker_name
            FROM bankroll_accounts ba
            WHERE ba.user_id = %s
            ORDER BY ba.bookmaker_name
        """, [user_id])
        bookmakers = [row['bookmaker_name'] for row in cursor.fetchall()]
        
        # Get unique markets
        cursor.execute("""
            SELECT DISTINCT market_key
            FROM bets
            WHERE user_id = %s
            AND market_key IS NOT NULL
            ORDER BY market_key
        """, [user_id])
        markets = [row['market_key'] for row in cursor.fetchall()]
        
        # Get unique sports
        cursor.execute("""
            SELECT DISTINCT sport
            FROM bets
            WHERE user_id = %s
            AND sport IS NOT NULL
            ORDER BY sport
        """, [user_id])
        sports = [row['sport'] for row in cursor.fetchall()]
        
        if cursor:
            cursor.close()
        
        return {
            "bookmakers": bookmakers,
            "markets": markets,
            "sports": sports,
            "statuses": ["pending", "won", "lost", "push", "cancelled"]
        }
        
    except Exception as e:
        print(f"❌ Filter Options Error: {str(e)}")
        traceback.print_exc()
        if cursor:
            try:
                cursor.close()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"Filter options failed: {str(e)}")

@router.get("/tax-report")
async def get_tax_report(
    user_id: int = Query(default=1),
    year: int = Query(default=2024),
    conn = Depends(get_db)
):
    """
    Generate tax report for IRS Schedule C.
    
    **Use Case:** Year-end tax preparation
    **Returns:** Comprehensive tax summary
    """
    cursor = None
    try:
        cursor = conn.cursor()
        
        # Get yearly summary
        cursor.execute("""
            SELECT 
                COUNT(*) as total_bets,
                COUNT(*) FILTER (WHERE status = 'won') as winning_bets,
                COUNT(*) FILTER (WHERE status = 'lost') as losing_bets,
                COUNT(*) FILTER (WHERE status = 'push') as push_bets,
                COALESCE(SUM(stake_amount), 0) as total_wagered,
                COALESCE(SUM(actual_payout) FILTER (WHERE status = 'won'), 0) as total_winnings,
                COALESCE(SUM(stake_amount) FILTER (WHERE status = 'lost'), 0) as total_losses,
                COALESCE(SUM(profit_loss) FILTER (WHERE status IN ('won', 'lost', 'push')), 0) as net_profit_loss
            FROM bets
            WHERE user_id = %s
            AND EXTRACT(YEAR FROM placed_at) = %s
        """, [user_id, year])
        
        summary = cursor.fetchone()
        
        # Get breakdown by bookmaker
        cursor.execute("""
            SELECT 
                ba.bookmaker_name,
                COUNT(*) as bets,
                COALESCE(SUM(b.stake_amount), 0) as wagered,
                COALESCE(SUM(b.actual_payout) FILTER (WHERE b.status = 'won'), 0) as winnings,
                COALESCE(SUM(b.stake_amount) FILTER (WHERE b.status = 'lost'), 0) as losses,
                COALESCE(SUM(b.profit_loss) FILTER (WHERE b.status IN ('won', 'lost', 'push')), 0) as net
            FROM bets b
            LEFT JOIN bankroll_accounts ba ON b.account_id = ba.account_id
            WHERE b.user_id = %s
            AND EXTRACT(YEAR FROM b.placed_at) = %s
            GROUP BY ba.bookmaker_name
            ORDER BY net DESC
        """, [user_id, year])
        
        by_bookmaker = cursor.fetchall()
        
        # Get breakdown by month
        cursor.execute("""
            SELECT 
                TO_CHAR(placed_at, 'Month') as month,
                EXTRACT(MONTH FROM placed_at) as month_num,
                COUNT(*) as bets,
                COALESCE(SUM(profit_loss) FILTER (WHERE status IN ('won', 'lost', 'push')), 0) as profit_loss
            FROM bets
            WHERE user_id = %s
            AND EXTRACT(YEAR FROM placed_at) = %s
            GROUP BY TO_CHAR(placed_at, 'Month'), EXTRACT(MONTH FROM placed_at)
            ORDER BY month_num
        """, [user_id, year])
        
        by_month = cursor.fetchall()
        
        # Get breakdown by sport
        cursor.execute("""
            SELECT 
                sport,
                COUNT(*) as bets,
                COALESCE(SUM(profit_loss) FILTER (WHERE status IN ('won', 'lost', 'push')), 0) as profit_loss,
                ROUND(
                    CASE 
                        WHEN COUNT(*) FILTER (WHERE status IN ('won', 'lost')) > 0 THEN
                            (COUNT(*) FILTER (WHERE status = 'won')::numeric / 
                             COUNT(*) FILTER (WHERE status IN ('won', 'lost')) * 100)
                        ELSE 0
                    END,
                1) as win_rate
            FROM bets
            WHERE user_id = %s
            AND EXTRACT(YEAR FROM placed_at) = %s
            AND sport IS NOT NULL
            GROUP BY sport
            ORDER BY profit_loss DESC
        """, [user_id, year])
        
        by_sport = cursor.fetchall()
        
        if cursor:
            cursor.close()
        
        # Format for IRS Schedule C
        return {
            "year": year,
            "summary": {
                "total_bets": summary['total_bets'],
                "winning_bets": summary['winning_bets'],
                "losing_bets": summary['losing_bets'],
                "push_bets": summary['push_bets'],
                "total_wagered": float(summary['total_wagered']),
                "total_winnings": float(summary['total_winnings']),
                "total_losses": float(summary['total_losses']),
                "net_profit_loss": float(summary['net_profit_loss']),
                "win_rate": round((summary['winning_bets'] / summary['total_bets'] * 100), 1) if summary['total_bets'] > 0 else 0
            },
            "schedule_c": {
                "line_1_gross_receipts": float(summary['total_winnings']),
                "line_28_total_expenses": float(summary['total_losses']),
                "line_31_net_profit": float(summary['net_profit_loss'])
            },
            "by_bookmaker": [
                {
                    "bookmaker": row['bookmaker_name'] or 'Unknown',
                    "bets": row['bets'],
                    "wagered": float(row['wagered']),
                    "winnings": float(row['winnings']),
                    "losses": float(row['losses']),
                    "net": float(row['net'])
                }
                for row in by_bookmaker
            ],
            "by_month": [
                {
                    "month": row['month'].strip(),
                    "month_num": row['month_num'],
                    "bets": row['bets'],
                    "profit_loss": float(row['profit_loss'])
                }
                for row in by_month
            ],
            "by_sport": [
                {
                    "sport": row['sport'],
                    "bets": row['bets'],
                    "profit_loss": float(row['profit_loss']),
                    "win_rate": float(row['win_rate'])
                }
                for row in by_sport
            ]
        }
        
    except Exception as e:
        print(f"❌ Tax Report Error: {str(e)}")
        traceback.print_exc()
        if cursor:
            try:
                cursor.close()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"Tax report failed: {str(e)}")


@router.get("/tax-report/pdf")
async def download_tax_report_pdf(
    user_id: int = Query(default=1),
    year: int = Query(default=2024),
    conn = Depends(get_db)
):
    """
    Download tax report as PDF.
    
    **Use Case:** Printable tax document
    **Returns:** PDF file download
    """
    # Get the tax report data
    tax_data = await get_tax_report(user_id, year, conn)
    
    # Create simple text-based PDF (we'll upgrade this later)
    from io import BytesIO
    
    # For now, create a formatted text file
    output = BytesIO()
    content = f"""
=====================================
BETTING TAX REPORT - {year}
=====================================

SUMMARY
-------
Total Bets: {tax_data['summary']['total_bets']}
Winning Bets: {tax_data['summary']['winning_bets']}
Losing Bets: {tax_data['summary']['losing_bets']}
Win Rate: {tax_data['summary']['win_rate']}%

Total Wagered: ${tax_data['summary']['total_wagered']:,.2f}
Total Winnings: ${tax_data['summary']['total_winnings']:,.2f}
Total Losses: ${tax_data['summary']['total_losses']:,.2f}
Net Profit/Loss: ${tax_data['summary']['net_profit_loss']:,.2f}

=====================================
IRS SCHEDULE C (Form 1040)
=====================================

Line 1 - Gross Receipts (Winnings):
    ${tax_data['schedule_c']['line_1_gross_receipts']:,.2f}

Line 28 - Total Expenses (Losses):
    ${tax_data['schedule_c']['line_28_total_expenses']:,.2f}

Line 31 - Net Profit (or Loss):
    ${tax_data['schedule_c']['line_31_net_profit']:,.2f}

=====================================
BREAKDOWN BY BOOKMAKER
=====================================

"""
    
    for bm in tax_data['by_bookmaker']:
        content += f"""
{bm['bookmaker']}:
    Bets: {bm['bets']}
    Wagered: ${bm['wagered']:,.2f}
    Winnings: ${bm['winnings']:,.2f}
    Losses: ${bm['losses']:,.2f}
    Net: ${bm['net']:,.2f}
"""
    
    content += """
=====================================
MONTHLY BREAKDOWN
=====================================

"""
    
    for month in tax_data['by_month']:
        content += f"{month['month']:12s} - {month['bets']:3d} bets - ${month['profit_loss']:,.2f}\n"
    
    content += """
=====================================
BREAKDOWN BY SPORT
=====================================

"""
    
    for sport in tax_data['by_sport']:
        content += f"""
{sport['sport']}:
    Bets: {sport['bets']}
    Win Rate: {sport['win_rate']}%
    Net: ${sport['profit_loss']:,.2f}
"""
    
    content += f"""

=====================================
Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}
=====================================

DISCLAIMER: This report is for informational purposes only.
Please consult with a tax professional for proper tax filing.
The IRS requires reporting of all gambling winnings and losses.
"""
    
    output.write(content.encode('utf-8'))
    output.seek(0)
    
    filename = f"tax_report_{year}_{datetime.now().strftime('%Y%m%d')}.txt"
    
    return StreamingResponse(
        output,
        media_type="text/plain",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

@router.get("/excel")
async def export_excel(
    user_id: int = Query(default=1),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    bookmaker: Optional[str] = Query(default=None),
    market: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    conn = Depends(get_db)
):
    """
    Export bets to Excel format with multiple sheets and formulas.
    
    **Use Case:** Advanced analysis with Excel
    **Returns:** Excel (.xlsx) file download
    """
    cursor = None
    try:
        cursor = conn.cursor()
        
        # Build WHERE clause (same as CSV)
        where_clauses = ["b.user_id = %s"]
        params = [user_id]
        
        if start_date:
            where_clauses.append("DATE(b.placed_at) >= %s")
            params.append(start_date)
        
        if end_date:
            where_clauses.append("DATE(b.placed_at) <= %s")
            params.append(end_date)
        
        if bookmaker:
            where_clauses.append("ba.bookmaker_name = %s")
            params.append(bookmaker)
        
        if market:
            where_clauses.append("b.market_key = %s")
            params.append(market)
        
        if status:
            where_clauses.append("b.status = %s")
            params.append(status)
        
        where_clause = " AND ".join(where_clauses)
        
        # Fetch all bets
        query = f"""
            SELECT 
                b.bet_id,
                TO_CHAR(b.placed_at, 'YYYY-MM-DD HH24:MI:SS') as placed_at,
                ba.bookmaker_name,
                b.bet_type,
                b.sport,
                b.market_key,
                b.bet_side,
                b.line_value,
                b.odds_american,
                b.stake_amount,
                b.potential_payout,
                b.status,
                b.actual_payout,
                b.profit_loss,
                TO_CHAR(b.settled_at, 'YYYY-MM-DD HH24:MI:SS') as settled_at,
                b.notes
            FROM bets b
            LEFT JOIN bankroll_accounts ba ON b.account_id = ba.account_id
            WHERE {where_clause}
            ORDER BY b.placed_at DESC
        """
        
        cursor.execute(query, params)
        bets = cursor.fetchall()
        
        # Get summary by bookmaker
        cursor.execute(f"""
            SELECT 
                ba.bookmaker_name,
                COUNT(*) as total_bets,
                COUNT(*) FILTER (WHERE b.status = 'won') as won_bets,
                COUNT(*) FILTER (WHERE b.status = 'lost') as lost_bets,
                COALESCE(SUM(b.stake_amount), 0) as total_staked,
                COALESCE(SUM(b.profit_loss), 0) as profit_loss,
                ROUND(
                    CASE 
                        WHEN COUNT(*) FILTER (WHERE b.status IN ('won', 'lost')) > 0 THEN
                            (COUNT(*) FILTER (WHERE b.status = 'won')::numeric / 
                             COUNT(*) FILTER (WHERE b.status IN ('won', 'lost')) * 100)
                        ELSE 0
                    END,
                1) as win_rate
            FROM bets b
            LEFT JOIN bankroll_accounts ba ON b.account_id = ba.account_id
            WHERE {where_clause}
            GROUP BY ba.bookmaker_name
            ORDER BY profit_loss DESC
        """, params)
        
        by_bookmaker = cursor.fetchall()
        
        # Get summary by market
        cursor.execute(f"""
            SELECT 
                b.market_key,
                COUNT(*) as total_bets,
                COUNT(*) FILTER (WHERE b.status = 'won') as won_bets,
                COALESCE(SUM(b.profit_loss), 0) as profit_loss,
                ROUND(
                    CASE 
                        WHEN COUNT(*) FILTER (WHERE b.status IN ('won', 'lost')) > 0 THEN
                            (COUNT(*) FILTER (WHERE b.status = 'won')::numeric / 
                             COUNT(*) FILTER (WHERE b.status IN ('won', 'lost')) * 100)
                        ELSE 0
                    END,
                1) as win_rate
            FROM bets b
            WHERE {where_clause}
            AND b.market_key IS NOT NULL
            GROUP BY b.market_key
            ORDER BY profit_loss DESC
        """, params)
        
        by_market = cursor.fetchall()
        
        if cursor:
            cursor.close()
        
        # Create Excel workbook
        wb = Workbook()
        
        # ===== SHEET 1: All Bets =====
        ws_bets = wb.active
        ws_bets.title = "All Bets"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = ['Bet ID', 'Date', 'Bookmaker', 'Type', 'Sport', 'Market', 
                   'Side', 'Line', 'Odds', 'Stake', 'Potential', 'Status', 
                   'Payout', 'Profit/Loss', 'Settled', 'Notes']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_bets.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_num, bet in enumerate(bets, 2):
            ws_bets.cell(row=row_num, column=1, value=bet['bet_id'])
            ws_bets.cell(row=row_num, column=2, value=bet['placed_at'])
            ws_bets.cell(row=row_num, column=3, value=bet['bookmaker_name'] or 'N/A')
            ws_bets.cell(row=row_num, column=4, value=bet['bet_type'])
            ws_bets.cell(row=row_num, column=5, value=bet['sport'])
            ws_bets.cell(row=row_num, column=6, value=bet['market_key'])
            ws_bets.cell(row=row_num, column=7, value=bet['bet_side'])
            ws_bets.cell(row=row_num, column=8, value=float(bet['line_value']) if bet['line_value'] else '')
            ws_bets.cell(row=row_num, column=9, value=bet['odds_american'])
            ws_bets.cell(row=row_num, column=10, value=float(bet['stake_amount']) if bet['stake_amount'] else 0)
            ws_bets.cell(row=row_num, column=11, value=float(bet['potential_payout']) if bet['potential_payout'] else 0)
            ws_bets.cell(row=row_num, column=12, value=bet['status'])
            ws_bets.cell(row=row_num, column=13, value=float(bet['actual_payout']) if bet['actual_payout'] else 0)
            ws_bets.cell(row=row_num, column=14, value=float(bet['profit_loss']) if bet['profit_loss'] else 0)
            ws_bets.cell(row=row_num, column=15, value=bet['settled_at'])
            ws_bets.cell(row=row_num, column=16, value=bet['notes'])
        
        # Auto-adjust column widths
        for column in ws_bets.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_bets.column_dimensions[column_letter].width = adjusted_width
        
        # ===== SHEET 2: Summary =====
        ws_summary = wb.create_sheet("Summary")
        
        # Title
        ws_summary['A1'] = 'BETTING SUMMARY'
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        # Summary stats
        total_bets = len(bets)
        won_bets = len([b for b in bets if b['status'] == 'won'])
        lost_bets = len([b for b in bets if b['status'] == 'lost'])
        total_staked = sum(float(b['stake_amount'] or 0) for b in bets)
        total_profit = sum(float(b['profit_loss'] or 0) for b in bets)
        win_rate = (won_bets / total_bets * 100) if total_bets > 0 else 0
        
        ws_summary['A3'] = 'Total Bets:'
        ws_summary['B3'] = total_bets
        ws_summary['A4'] = 'Winning Bets:'
        ws_summary['B4'] = won_bets
        ws_summary['A5'] = 'Losing Bets:'
        ws_summary['B5'] = lost_bets
        ws_summary['A6'] = 'Win Rate:'
        ws_summary['B6'] = f"{win_rate:.1f}%"
        ws_summary['A7'] = 'Total Staked:'
        ws_summary['B7'] = total_staked
        ws_summary['B7'].number_format = '$#,##0.00'
        ws_summary['A8'] = 'Total Profit/Loss:'
        ws_summary['B8'] = total_profit
        ws_summary['B8'].number_format = '$#,##0.00'
        
        # Style
        for row in range(3, 9):
            ws_summary[f'A{row}'].font = Font(bold=True)
        
        # ===== SHEET 3: By Bookmaker =====
        ws_bookmaker = wb.create_sheet("By Bookmaker")
        
        headers_bm = ['Bookmaker', 'Total Bets', 'Won', 'Lost', 'Staked', 'Profit/Loss', 'Win Rate %']
        for col_num, header in enumerate(headers_bm, 1):
            cell = ws_bookmaker.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        for row_num, bm in enumerate(by_bookmaker, 2):
            ws_bookmaker.cell(row=row_num, column=1, value=bm['bookmaker_name'] or 'Unknown')
            ws_bookmaker.cell(row=row_num, column=2, value=bm['total_bets'])
            ws_bookmaker.cell(row=row_num, column=3, value=bm['won_bets'])
            ws_bookmaker.cell(row=row_num, column=4, value=bm['lost_bets'])
            ws_bookmaker.cell(row=row_num, column=5, value=float(bm['total_staked']))
            ws_bookmaker.cell(row=row_num, column=6, value=float(bm['profit_loss']))
            ws_bookmaker.cell(row=row_num, column=7, value=float(bm['win_rate']))
        
        # ===== SHEET 4: By Market =====
        ws_market = wb.create_sheet("By Market")
        
        headers_mk = ['Market', 'Total Bets', 'Won', 'Profit/Loss', 'Win Rate %']
        for col_num, header in enumerate(headers_mk, 1):
            cell = ws_market.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        for row_num, mk in enumerate(by_market, 2):
            ws_market.cell(row=row_num, column=1, value=mk['market_key'])
            ws_market.cell(row=row_num, column=2, value=mk['total_bets'])
            ws_market.cell(row=row_num, column=3, value=mk['won_bets'])
            ws_market.cell(row=row_num, column=4, value=float(mk['profit_loss']))
            ws_market.cell(row=row_num, column=5, value=float(mk['win_rate']))
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"bets_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        print(f"❌ Excel Export Error: {str(e)}")
        traceback.print_exc()
        if cursor:
            try:
                cursor.close()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"Excel export failed: {str(e)}")

@router.get("/pdf-report")
async def export_pdf_report(
    user_id: int = Query(default=1),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    report_type: str = Query(default="summary"),  # 'summary', 'detailed', 'tax'
    conn = Depends(get_db)
):
    """
    Export betting report as PDF.
    
    **Use Case:** Professional formatted reports
    **Returns:** PDF file download
    """
    cursor = None
    try:
        cursor = conn.cursor()
        
        # Build WHERE clause
        where_clauses = ["b.user_id = %s"]
        params = [user_id]
        
        if start_date:
            where_clauses.append("DATE(b.placed_at) >= %s")
            params.append(start_date)
        
        if end_date:
            where_clauses.append("DATE(b.placed_at) <= %s")
            params.append(end_date)
        
        where_clause = " AND ".join(where_clauses)
        
        # Get summary data
        cursor.execute(f"""
            SELECT 
                COUNT(*) as total_bets,
                COUNT(*) FILTER (WHERE status = 'won') as won_bets,
                COUNT(*) FILTER (WHERE status = 'lost') as lost_bets,
                COUNT(*) FILTER (WHERE status = 'push') as push_bets,
                COUNT(*) FILTER (WHERE status = 'pending') as pending_bets,
                COALESCE(SUM(stake_amount), 0) as total_staked,
                COALESCE(SUM(profit_loss) FILTER (WHERE status IN ('won', 'lost', 'push')), 0) as total_profit,
                ROUND(
                    CASE 
                        WHEN COUNT(*) FILTER (WHERE status IN ('won', 'lost')) > 0 THEN
                            (COUNT(*) FILTER (WHERE status = 'won')::numeric / 
                             COUNT(*) FILTER (WHERE status IN ('won', 'lost')) * 100)
                        ELSE 0
                    END,
                1) as win_rate,
                MIN(DATE(placed_at)) as earliest_bet,
                MAX(DATE(placed_at)) as latest_bet
            FROM bets b
            WHERE {where_clause}
        """, params)
        
        summary = cursor.fetchone()
        
        # Get by bookmaker
        cursor.execute(f"""
            SELECT 
                ba.bookmaker_name,
                COUNT(*) as bets,
                COALESCE(SUM(b.profit_loss), 0) as profit_loss,
                ROUND(
                    CASE 
                        WHEN COUNT(*) FILTER (WHERE b.status IN ('won', 'lost')) > 0 THEN
                            (COUNT(*) FILTER (WHERE b.status = 'won')::numeric / 
                             COUNT(*) FILTER (WHERE b.status IN ('won', 'lost')) * 100)
                        ELSE 0
                    END,
                1) as win_rate
            FROM bets b
            LEFT JOIN bankroll_accounts ba ON b.account_id = ba.account_id
            WHERE {where_clause}
            GROUP BY ba.bookmaker_name
            ORDER BY profit_loss DESC
        """, params)
        
        by_bookmaker = cursor.fetchall()
        
        if cursor:
            cursor.close()
        
        # Create PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        date_range = ""
        if start_date and end_date:
            date_range = f"{start_date} to {end_date}"
        elif start_date:
            date_range = f"From {start_date}"
        elif end_date:
            date_range = f"Until {end_date}"
        else:
            date_range = "All Time"
        
        story.append(Paragraph("Betting Performance Report", title_style))
        story.append(Paragraph(f"<i>{date_range}</i>", styles['Normal']))
        story.append(Spacer(1, 0.5*inch))
        
        # Summary Section
        story.append(Paragraph("<b>Summary Statistics</b>", styles['Heading2']))
        story.append(Spacer(1, 0.2*inch))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Bets', str(summary['total_bets'])],
            ['Winning Bets', str(summary['won_bets'])],
            ['Losing Bets', str(summary['lost_bets'])],
            ['Push Bets', str(summary['push_bets'])],
            ['Pending Bets', str(summary['pending_bets'])],
            ['Win Rate', f"{summary['win_rate']}%"],
            ['Total Staked', f"${float(summary['total_staked']):,.2f}"],
            ['Total Profit/Loss', f"${float(summary['total_profit']):,.2f}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 0.5*inch))
        
        # Performance by Bookmaker
        if by_bookmaker:
            story.append(Paragraph("<b>Performance by Bookmaker</b>", styles['Heading2']))
            story.append(Spacer(1, 0.2*inch))
            
            bookmaker_data = [['Bookmaker', 'Bets', 'Win Rate', 'Profit/Loss']]
            for bm in by_bookmaker:
                bookmaker_data.append([
                    bm['bookmaker_name'] or 'Unknown',
                    str(bm['bets']),
                    f"{bm['win_rate']}%",
                    f"${float(bm['profit_loss']):,.2f}"
                ])
            
            bookmaker_table = Table(bookmaker_data, colWidths=[2*inch, 1*inch, 1*inch, 1.5*inch])
            bookmaker_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(bookmaker_table)
        
        # Footer
        story.append(Spacer(1, 1*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        story.append(Paragraph(
            f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}<br/>"
            "SmartLine Bankroll Manager",
            footer_style
        ))
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        filename = f"betting_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        print(f"❌ PDF Export Error: {str(e)}")
        traceback.print_exc()
        if cursor:
            try:
                cursor.close()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"PDF export failed: {str(e)}")
